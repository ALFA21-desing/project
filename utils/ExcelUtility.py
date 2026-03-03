"""
ExcelUtility - Utility class for reading test data from Excel files
Supports reading login credentials and other test data
"""

import openpyxl
import csv
import os


class ExcelUtility:
    """Utility class to read test data from Excel and CSV files"""
    
    @staticmethod
    def read_excel(file_path, sheet_name=None):
        """
        Read data from an Excel file
        
        Args:
            file_path (str): Path to the Excel file
            sheet_name (str): Name of the sheet to read (default: first sheet)
            
        Returns:
            list: List of dictionaries containing row data
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path)
        
        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Read data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Convert None cell values to empty strings for consistency with CSV reading
            row_dict = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
            data.append(row_dict)
        
        workbook.close()
        return data
    
    @staticmethod
    def read_csv(file_path):
        """
        Read data from a CSV file
        
        Args:
            file_path (str): Path to the CSV file
            
        Returns:
            list: List of dictionaries containing row data
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"CSV file not found: {file_path}")
        
        data = []
        with open(file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            for row in csv_reader:
                data.append(row)
        
        return data
    
    @staticmethod
    def write_excel(file_path, data, sheet_name="Sheet1"):
        """
        Write data to an Excel file
        
        Args:
            file_path (str): Path to save the Excel file
            data (list): List of dictionaries to write
            sheet_name (str): Name of the sheet
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        if data:
            # Write headers
            headers = list(data[0].keys())
            sheet.append(headers)
            
            # Write data rows
            for row_dict in data:
                row = [row_dict[header] for header in headers]
                sheet.append(row)
        
        workbook.save(file_path)
        workbook.close()
