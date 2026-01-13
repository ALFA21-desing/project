#!/usr/bin/env python3
"""Extract first sheet from available login_credentials XLSX to CSV.

Usage:
  python scripts/extract_xlsx_to_csv.py

Creates: test_data/login_credentials.csv
"""
from pathlib import Path
import csv

try:
    from openpyxl import load_workbook
except Exception as e:
    print('openpyxl is required. Install with: pip install openpyxl')
    raise

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / 'test_data'
CANDIDATES = [
    DATA_DIR / 'login_credentials.csv.xlsx',
    DATA_DIR / 'login_credentialsT2.csv.xlsx',
    DATA_DIR / 'login_credentials.xlsx'
]

def find_file():
    for p in CANDIDATES:
        if p.exists():
            return p
    return None

def convert(xlsx_path, csv_path):
    wb = load_workbook(filename=str(xlsx_path), read_only=True)
    ws = wb.active
    with csv_path.open('w', newline='', encoding='utf-8') as fh:
        writer = csv.writer(fh)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else str(v) for v in row])

def main():
    src = find_file()
    if not src:
        print('No candidate XLSX found in test_data/.')
        return
    dest = DATA_DIR / 'login_credentials.csv'
    print(f'Converting {src} -> {dest}')
    convert(src, dest)
    print('Done.')

if __name__ == '__main__':
    main()
